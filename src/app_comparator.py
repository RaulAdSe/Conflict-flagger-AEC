import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

# --- PARSER BC3 ---
class BC3Parser:
    def __init__(self, filepath):
        self.filepath = filepath

    def parse(self):
        """Devuelve {codigo: {'quantity': float, 'unit': str, 'desc': str}}"""
        data = {}
        try:
            try:
                with open(self.filepath, 'r', encoding='latin-1') as f: content = f.read()
            except:
                with open(self.filepath, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
            
            lines = content.split('\n')
            
            for line in lines:
                line = line.strip()
                if line.startswith('~C|'):
                    parts = line.split('|')
                    if len(parts) >= 4:
                        code = parts[1].strip().rstrip('#')
                        unit = parts[2].strip()
                        desc = parts[3].strip()
                        data[code] = {'desc': desc, 'unit': unit, 'quantity': 0.0}

            for line in lines:
                line = line.strip()
                if line.startswith('~D|'):
                    parts = line.split('|')
                    if len(parts) >= 3:
                        items = parts[2].split('\\')
                        i = 0
                        while i < len(items):
                            child_code = items[i].strip()
                            if child_code and i + 2 < len(items):
                                try:
                                    qty = float(items[i + 2].strip())
                                    if child_code in data: data[child_code]['quantity'] += qty
                                except: pass
                            i += 3
            return data
        except Exception as e:
            messagebox.showerror("Error BC3", f"Error leyendo BC3: {str(e)}")
            return {}


# --- PARSER IFC CON DESCRIPCIONES ---
class IFCParser:
    def __init__(self, filepath):
        self.filepath = filepath

    def parse(self):
        """
        Devuelve:
        - codes: {codigo: count}
        - descriptions: {codigo: descripcion}
        """
        codes = {}
        descriptions = {}
        
        try:
            with open(self.filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            lines = content.split('\n')
            id_map = {}
            for line in lines:
                if line.startswith('#'):
                    idx = line.find('=')
                    if idx != -1:
                        id_map[line[:idx].strip()] = line[idx+1:].strip()
            
            code_to_entities = {}
            
            for oid, cont in id_map.items():
                if 'IFC' in cont:
                    all_values = re.findall(r"'([^']*)'", cont)
                    code = None
                    desc = None
                    
                    for val in all_values:
                        if val.isdigit() and 3 <= len(val) <= 10:
                            code = val
                        elif ':' in val and len(val) > 5:
                            desc = val
                    
                    if code:
                        if code not in code_to_entities:
                            code_to_entities[code] = []
                        code_to_entities[code].append(oid)
                        
                        if desc and code not in descriptions:
                            # Limpiar descripción IFC (quitar \X\ED etc)
                            desc_clean = re.sub(r'\\X\\[0-9A-F]{2}', '', desc)
                            descriptions[code] = desc_clean
            
            # Contar instancias
            for code, entities in code_to_entities.items():
                count = 0
                for oid in entities:
                    for rel_cont in id_map.values():
                        if 'IFCRELDEFINES' in rel_cont and oid in rel_cont:
                            match_list = re.search(r"\(\s*(#[0-9,\s#]+)\s*\)", rel_cont)
                            if match_list:
                                instances = match_list.group(1).replace(' ', '').split(',')
                                count += len(instances)
                                break
                codes[code] = max(1, count)
            
            return codes, descriptions

        except Exception as e:
            messagebox.showerror("Error IFC", f"Error leyendo IFC: {str(e)}")
            return {}, {}


# --- UTILIDADES DE MATCHING ---
def normalize_desc(desc):
    """Normaliza descripción para comparación"""
    if not desc:
        return ""
    # Quitar caracteres especiales, convertir a minúsculas
    desc = re.sub(r'[^a-zA-Z0-9áéíóúñÁÉÍÓÚÑ\s]', ' ', desc.lower())
    # Quitar espacios múltiples
    desc = ' '.join(desc.split())
    return desc

def calc_similarity(desc1, desc2):
    """Calcula similitud entre dos descripciones (0-1)"""
    words1 = set(normalize_desc(desc1).split())
    words2 = set(normalize_desc(desc2).split())
    
    if not words1 or not words2:
        return 0
    
    # Quitar palabras muy comunes
    stopwords = {'de', 'la', 'el', 'en', 'con', 'para', 'por', 'a', 'y', 'o', 'mm', 'cm', 'm', 'm2', 'm3'}
    words1 = words1 - stopwords
    words2 = words2 - stopwords
    
    if not words1 or not words2:
        return 0
    
    intersection = len(words1 & words2)
    union = len(words1 | words2)
    
    return intersection / union if union > 0 else 0


# --- APLICACIÓN PRINCIPAL ---
class ComparadorBIMApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparador BC3 vs IFC - Con Matching")
        self.root.geometry("600x250")

        self.path_bc3 = tk.StringVar()
        self.path_ifc = tk.StringVar()

        frame = tk.Frame(root, padx=20, pady=20)
        frame.pack(expand=True, fill='both')

        tk.Label(frame, text="1. Archivo BC3:").grid(row=0, column=0, sticky='w')
        tk.Entry(frame, textvariable=self.path_bc3, width=50).grid(row=0, column=1, padx=5)
        tk.Button(frame, text="...", command=lambda: self.browse(self.path_bc3, "*.bc3")).grid(row=0, column=2)

        tk.Label(frame, text="2. Archivo IFC:").grid(row=1, column=0, sticky='w', pady=10)
        tk.Entry(frame, textvariable=self.path_ifc, width=50).grid(row=1, column=1, padx=5, pady=10)
        tk.Button(frame, text="...", command=lambda: self.browse(self.path_ifc, "*.ifc")).grid(row=1, column=2)

        tk.Button(frame, text="COMPARAR", command=self.run_comparison, bg="#4CAF50", fg="white", 
                  height=2, font=('Arial', 10, 'bold')).grid(row=2, column=0, columnspan=3, pady=20, sticky='ew')

    def browse(self, var, filetype):
        f = filedialog.askopenfilename(filetypes=[("Archivos", filetype)])
        if f: var.set(f)

    def is_ignored_item(self, code, desc):
        """Ignora elementos que no son partidas de obra comparables"""
        ignore_terms = [
            "información", "project info", "plano", "sheet", "vista", "view",
            "zona de", "climatización", "topografía", "habitaciones", "rooms",
            "áreas", "areas", "ocupacion", "sup.libre", "sup.construida",
            "almacén", "salón", "cocina", "aseo", "archivo", "circulación",
            "área de trabajo", "sala de reuniones", "dep. limpieza",
            "aseos femeninos", "aseos masculinos",
            "aberturas", "hueco", "opening", "void", "corte", "líneas", "lines",
            "materiales", "materials", "tubería", "pipe", "segmentos",
            "system panel", "empty panel"
        ]
        text_check = (str(desc) + " " + str(code)).lower()
        return any(t in text_check for t in ignore_terms)

    def find_matching_code(self, bc3_code, bc3_desc, ifc_descriptions, threshold=0.5):
        """Busca en IFC un código con descripción similar. Si hay empate, prefiere código numéricamente cercano."""
        matches = []
        
        for ifc_code, ifc_desc in ifc_descriptions.items():
            score = calc_similarity(bc3_desc, ifc_desc)
            if score >= threshold:
                matches.append((ifc_code, score))
        
        if not matches:
            return None, 0
        
        # Si solo hay un match, devolverlo
        if len(matches) == 1:
            return matches[0]
        
        # Si hay múltiples matches con el mismo score, desempatar por cercanía numérica
        max_score = max(m[1] for m in matches)
        top_matches = [m for m in matches if m[1] == max_score]
        
        if len(top_matches) == 1:
            return top_matches[0]
        
        # Desempatar por cercanía numérica al código BC3
        try:
            bc3_num = int(bc3_code)
            best = min(top_matches, key=lambda m: abs(int(m[0]) - bc3_num))
            return best
        except:
            return top_matches[0]

    def run_comparison(self):
        bc3_path = self.path_bc3.get()
        ifc_path = self.path_ifc.get()

        if not bc3_path or not ifc_path:
            messagebox.showwarning("Aviso", "Selecciona ambos archivos.")
            return

        # Parsear archivos
        bc3_data = BC3Parser(bc3_path).parse()
        ifc_counts, ifc_descriptions = IFCParser(ifc_path).parse()

        discrepancias = []
        coincidencias = []

        for code, info in bc3_data.items():
            if not code or len(code) > 40: continue
            if not info['unit']: continue
            
            desc = info['desc']
            unit = info['unit']
            qty_bc3 = info['quantity']
            
            if self.is_ignored_item(code, desc): continue
            if qty_bc3 == 0: continue
            
            code_in_ifc = code in ifc_counts
            ifc_count = ifc_counts.get(code, 0)
            
            status = "OK"
            msg = ""
            codigo_ifc = code if code_in_ifc else "-"

            if not code_in_ifc:
                # Buscar por descripción similar
                matching_code, score = self.find_matching_code(code, desc, ifc_descriptions)
                
                if matching_code:
                    status = "CODIGO_DIFERENTE"
                    codigo_ifc = matching_code
                    ifc_count = ifc_counts.get(matching_code, 0)
                    msg = f"BC3 usa '{code}' pero IFC usa '{matching_code}'"
                    
                    # También verificar cantidad si es unidad de conteo
                    if unit.lower() in ['u', 'ud', 'un', 'pza', 'ut']:
                        if abs(qty_bc3 - ifc_count) > 0.5:
                            msg += f" | Cantidad también difiere: BC3={int(qty_bc3)} vs IFC={ifc_count}"
                else:
                    status = "CODIGO_NO_EN_IFC"
                    msg = f"Código '{code}' no encontrado en IFC (sin match por descripción)"
            else:
                # Código existe en ambos - verificar cantidad
                is_count_unit = unit.lower() in ['u', 'ud', 'un', 'pza', 'ut']
                if is_count_unit:
                    if abs(qty_bc3 - ifc_count) > 0.5:
                        status = "CANTIDAD_DIFERENTE"
                        msg = f"BC3: {int(qty_bc3)} {unit} | IFC: {ifc_count} instancias"
                    else:
                        msg = f"OK: {int(qty_bc3)} {unit}"
                else:
                    msg = f"Código presente en ambos"

            row = {
                'CODIGO_BC3': code,
                'CODIGO_IFC': codigo_ifc,
                'DESCRIPCION': desc,
                'UNIDAD': unit,
                'CANT_BC3': qty_bc3,
                'CANT_IFC': ifc_count if codigo_ifc != "-" else "-",
                'ESTADO': status,
                'DETALLE': msg
            }

            if status == "OK":
                coincidencias.append(row)
            else:
                discrepancias.append(row)

        self.export_results(discrepancias, coincidencias, ifc_path)

    def export_results(self, discrepancias, coincidencias, ifc_path):
        try:
            base_name = os.path.splitext(os.path.basename(ifc_path))[0]
            output_file = f"{base_name}_comparacion.xlsx"
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                orange_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
                success_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                white_font = Font(color="FFFFFF", bold=True)
                bold_font = Font(bold=True)

                if not discrepancias:
                    df_success = pd.DataFrame({'RESULTADO': ["✓ NO EXISTE NINGUNA DISCREPANCIA"]})
                    df_success.to_excel(writer, sheet_name='Resumen', index=False)
                    ws = writer.sheets['Resumen']
                    ws['A2'].fill = success_fill
                    ws['A2'].font = white_font
                    ws.column_dimensions['A'].width = 50
                else:
                    df_disc = pd.DataFrame(discrepancias)
                    priority = {'CODIGO_DIFERENTE': 0, 'CANTIDAD_DIFERENTE': 1, 'CODIGO_NO_EN_IFC': 2}
                    df_disc['_sort'] = df_disc['ESTADO'].map(lambda x: priority.get(x, 99))
                    df_disc = df_disc.sort_values('_sort').drop('_sort', axis=1)
                    
                    df_disc.to_excel(writer, sheet_name='Discrepancias', index=False)
                    ws = writer.sheets['Discrepancias']
                    
                    for row_idx, row_data in enumerate(df_disc.itertuples(), start=2):
                        st = row_data.ESTADO
                        if st == 'CODIGO_DIFERENTE':
                            fill = orange_fill
                        elif st == 'CANTIDAD_DIFERENTE':
                            fill = red_fill
                        else:
                            fill = yellow_fill
                        for col_idx in range(1, len(df_disc.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                    
                    for col_idx in range(1, len(df_disc.columns) + 1):
                        ws.cell(row=1, column=col_idx).font = bold_font
                    
                    ws.column_dimensions['A'].width = 12
                    ws.column_dimensions['B'].width = 12
                    ws.column_dimensions['C'].width = 50
                    ws.column_dimensions['H'].width = 55

                if coincidencias:
                    df_match = pd.DataFrame(coincidencias)
                    df_match.to_excel(writer, sheet_name='Coincidencias', index=False)
                    ws_match = writer.sheets['Coincidencias']
                    for row_idx in range(2, len(df_match) + 2):
                        for col_idx in range(1, len(df_match.columns) + 1):
                            ws_match.cell(row=row_idx, column=col_idx).fill = green_fill

            msg = f"Reporte: {output_file}\n\n"
            msg += f"🔴 Discrepancias: {len(discrepancias)}\n"
            msg += f"🟢 Coincidencias: {len(coincidencias)}"
            
            messagebox.showinfo("Completado", msg)
            if os.name == 'nt':
                os.startfile(output_file)

        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = ComparadorBIMApp(root)
    root.mainloop()