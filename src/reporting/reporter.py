"""
Reporter for generating Excel reports with conflict visualization.

Generates color-coded Excel reports showing differences between IFC and BC3.
Reports are in Spanish for AEC professionals.

PHASE SUPPORT:
==============
The generate_report() method accepts an optional PhaseConfig that controls:
- Which sheets to generate
- Whether to include summary sheet
- Whether to show OK matches

This allows the same Reporter to produce different outputs for quick checks
vs comprehensive analysis without code duplication.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, TYPE_CHECKING
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.matching.matcher import MatchResult, MatchedPair, MatchStatus
from src.comparison.comparator import (
    ComparisonResult, Conflict, ConflictType, ConflictSeverity
)

# Import PhaseConfig only for type checking to avoid circular imports
if TYPE_CHECKING:
    from src.phases.config import PhaseConfig


# Spanish translations for conflict types and messages
TRANSLATIONS = {
    # Conflict types - Phase 1
    "code_mismatch": "Código diferente",
    "quantity_mismatch": "Cantidad diferente",
    "unit_mismatch": "Unidad diferente",
    "missing_in_bc3": "Sin presupuestar",
    "missing_in_ifc": "Sin modelar",

    # Conflict types - Phase 2
    "property_mismatch": "Diferencia de valor",
    "property_missing_ifc": "Propiedad falta en IFC",
    "property_missing_bc3": "Propiedad falta en BC3",
    "name_mismatch": "Nombre diferente",
    "type_mismatch": "Diferencia de tipo",

    # Severity levels
    "error": "ERROR",
    "warning": "AVISO",
    "info": "INFO",

    # Match methods
    "guid": "Por GUID",
    "tag": "Por Tag/ID",
    "name": "Por nombre",
    "description": "Por descripción",
    "type_name": "Por tipo",

    # Common messages
    "Element exists in IFC but not in BC3 budget": "Elemento existe en el modelo IFC pero no en el presupuesto BC3",
    "Element exists in BC3 but not in IFC model": "Elemento existe en el presupuesto BC3 pero no en el modelo IFC",
}


def translate(text: str) -> str:
    """Translate text to Spanish if translation exists."""
    return TRANSLATIONS.get(text, text)


@dataclass
class ReportConfig:
    """Configuration for report generation."""

    # Colors (RGB hex)
    color_error: str = "FF6B6B"  # Soft red - quantity mismatch
    color_warning: str = "FFE66D"  # Soft yellow - missing items
    color_ok: str = "7ED957"  # Soft green - correct
    color_info: str = "74C0FC"  # Soft blue - informational
    color_header: str = "2E86AB"  # Professional blue
    color_code_mismatch: str = "FFA500"  # Orange - code mismatch (matched by description)

    # Display options
    show_ok_matches: bool = True
    show_info_conflicts: bool = False
    max_rows: int = 10000


class Reporter:
    """Generates Excel reports for IFC-BC3 comparison results."""

    def __init__(self, config: Optional[ReportConfig] = None):
        """
        Initialize the reporter.

        Args:
            config: Report configuration options
        """
        self.config = config or ReportConfig()
        self._thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

    def generate_report(
        self,
        match_result: MatchResult,
        comparison_result: ComparisonResult,
        output_path: str | Path,
        include_summary: bool = True,
        phase_config: 'PhaseConfig' = None
    ) -> Path:
        """
        Generate an Excel report.

        Args:
            match_result: Result from the Matcher
            comparison_result: Result from the Comparator
            output_path: Path for the output Excel file
            include_summary: Whether to include a summary sheet (can be overridden by phase_config)
            phase_config: Optional phase configuration that controls which sheets to generate.
                         If None, generates all sheets (full analysis mode).

        Returns:
            Path to the generated report

        Phase Support:
            - QUICK_CHECK: Generates only Discrepàncies sheet
            - FULL_ANALYSIS: Generates all sheets (Resumen, Discrepancias, etc.)
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Determine which sheets to generate based on phase config
        if phase_config is not None:
            sheets_to_generate = set(phase_config.sheets)
            include_summary = phase_config.include_summary
        else:
            # Default to all sheets (full analysis)
            sheets_to_generate = {"Resumen", "Discrepancias", "Elementos Emparejados",
                                  "Sin Presupuestar", "Sin Modelar"}

        # Create sheets based on configuration
        if include_summary and ("Resumen" in sheets_to_generate or "Resum" in sheets_to_generate):
            self._create_summary_sheet(wb, match_result, comparison_result)

        # Discrepancies sheet (always included in any phase)
        if any(s in sheets_to_generate for s in ["Discrepancias", "Discrepàncies"]):
            self._create_conflicts_sheet(wb, comparison_result)

        # Matched elements sheet (supports both names)
        if "Elementos Emparejados" in sheets_to_generate or "Coincidencias" in sheets_to_generate:
            sheet_name = "Coincidencias" if "Coincidencias" in sheets_to_generate else "Elementos Emparejados"
            self._create_matches_sheet(wb, match_result, comparison_result, sheet_name)

        # Missing in budget sheet
        if "Sin Presupuestar" in sheets_to_generate:
            self._create_missing_sheet(wb, match_result, "Sin Presupuestar", MatchStatus.IFC_ONLY)

        # Missing in model sheet
        if "Sin Modelar" in sheets_to_generate:
            self._create_missing_sheet(wb, match_result, "Sin Modelar", MatchStatus.BC3_ONLY)

        # Elements summary sheet (always last)
        self._create_elements_summary_sheet(wb, match_result, comparison_result)

        wb.save(output_path)
        return output_path

    def _apply_header_style(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color=self.config.color_header,
            end_color=self.config.color_header,
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self._thin_border

    def _apply_cell_style(self, cell, fill_color: str = None) -> None:
        """Apply standard cell styling."""
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = self._thin_border
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            )

    def _create_summary_sheet(
        self,
        wb: Workbook,
        match_result: MatchResult,
        comparison_result: ComparisonResult
    ) -> None:
        """Create the summary sheet."""
        ws = wb.create_sheet("Resumen", 0)

        # Title
        ws["A1"] = "INFORME DE COMPARACION IFC - BC3"
        ws["A1"].font = Font(bold=True, size=18, color="2E86AB")
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 30

        # Subtitle with date
        ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells("A2:D2")

        # Summary box
        row = 4
        ws[f"A{row}"] = "RESUMEN EJECUTIVO"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        summary = comparison_result.summary()
        total_issues = summary["errors"] + summary["warnings"]

        if total_issues == 0:
            status_text = "Sin discrepancias detectadas"
            status_color = self.config.color_ok
        elif summary["errors"] > 0:
            status_text = f"{summary['errors']} errores y {summary['warnings']} avisos detectados"
            status_color = self.config.color_error
        else:
            status_text = f"{summary['warnings']} avisos detectados"
            status_color = self.config.color_warning

        ws[f"A{row}"] = status_text
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        ws.merge_cells(f"A{row}:D{row}")

        # Match statistics section
        row += 2
        ws[f"A{row}"] = "ESTADISTICAS DE EMPAREJAMIENTO"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:D{row}")

        stats = [
            ("Tipos en modelo IFC", match_result.total_ifc_types, "Familias/tipos encontrados en el archivo IFC"),
            ("Partidas en presupuesto BC3", match_result.total_bc3_elements, "Elementos encontrados en el archivo BC3"),
            ("Emparejados correctamente", len(match_result.matched), "Elementos que coinciden entre modelo y presupuesto"),
            ("Solo en IFC (sin presupuestar)", len(match_result.ifc_only), "Elementos modelados pero no incluidos en presupuesto"),
            ("Solo en BC3 (sin modelar)", len(match_result.bc3_only), "Partidas presupuestadas pero no modeladas"),
            ("Tasa de emparejamiento", f"{match_result.match_rate:.1f}%", "Porcentaje de elementos emparejados"),
        ]

        row += 1
        headers = ["Concepto", "Valor", "Descripcion"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)

        row += 1
        for label, value, description in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=description)
            for col in range(1, 4):
                self._apply_cell_style(ws.cell(row=row, column=col))
            row += 1

        # Comparison results section
        row += 1
        ws[f"A{row}"] = "RESULTADOS DE LA COMPARACION"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:D{row}")

        comp_stats = [
            ("Total de discrepancias", summary["total_conflicts"], self.config.color_info, "Numero total de diferencias encontradas"),
            ("Errores (diferencias)", summary["errors"], self.config.color_error if summary["errors"] > 0 else None, "Valores que no coinciden entre modelo y presupuesto"),
            ("Avisos (elementos faltantes)", summary["warnings"], self.config.color_warning if summary["warnings"] > 0 else None, "Elementos que existen en uno pero no en otro"),
            # Phase 1 specific
            ("Codigos diferentes", summary.get("code_mismatches", 0), self.config.color_code_mismatch if summary.get("code_mismatches", 0) > 0 else None, "Elementos emparejados por descripcion con codigos diferentes"),
            ("Cantidades diferentes", summary.get("quantity_mismatches", 0), self.config.color_error if summary.get("quantity_mismatches", 0) > 0 else None, "Cantidades que no coinciden"),
            # Phase 2 specific
            ("Propiedades diferentes", summary.get("property_mismatches", 0), None, "Propiedades con valores diferentes"),
        ]

        row += 1
        headers = ["Tipo", "Cantidad", "Descripcion"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)

        row += 1
        for label, value, color, description in comp_stats:
            ws.cell(row=row, column=1, value=label)
            cell_value = ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=description)
            for col in range(1, 4):
                self._apply_cell_style(ws.cell(row=row, column=col), color if col == 2 else None)
            row += 1

        # Legend
        row += 2
        ws[f"A{row}"] = "LEYENDA DE COLORES"
        ws[f"A{row}"].font = Font(bold=True, size=11)

        row += 1
        legend = [
            (self.config.color_error, "Rojo", "Error - Cantidad diferente, requiere atencion"),
            (self.config.color_code_mismatch, "Naranja", "Codigo diferente - Emparejado por descripcion"),
            (self.config.color_warning, "Amarillo", "Aviso - Elemento faltante, revisar"),
            (self.config.color_ok, "Verde", "Correcto - Sin problemas"),
        ]

        for color, name, desc in legend:
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=desc)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 15

    def _create_conflicts_sheet(
        self,
        wb: Workbook,
        comparison_result: ComparisonResult
    ) -> None:
        """Create the conflicts detail sheet."""
        ws = wb.create_sheet("Discrepancias")

        # Headers in Spanish
        headers = [
            ("Codigo", 15),
            ("Elemento", 30),
            ("Tipo de Discrepancia", 25),
            ("Gravedad", 12),
            ("Propiedad", 20),
            ("Valor IFC (Modelo)", 20),
            ("Valor BC3 (Presupuesto)", 20),
            ("Descripcion del Problema", 50)
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 25

        # Filter conflicts
        # Exclude MISSING_IN_BC3 and MISSING_IN_IFC - these have dedicated sheets
        conflicts = [c for c in comparison_result.conflicts
                     if c.conflict_type not in (ConflictType.MISSING_IN_BC3, ConflictType.MISSING_IN_IFC)]
        if not self.config.show_info_conflicts:
            conflicts = [c for c in conflicts if c.severity != ConflictSeverity.INFO]

        # Sort conflicts by color: red first, then orange, then yellow (Issue #9)
        # Red = ERROR (not code mismatch), Orange = CODE_MISMATCH, Yellow = WARNING
        def conflict_sort_key(c):
            # Determine color-based order
            if c.severity == ConflictSeverity.ERROR and c.conflict_type != ConflictType.CODE_MISMATCH:
                color_order = 0  # Red first
            elif c.conflict_type == ConflictType.CODE_MISMATCH:
                color_order = 1  # Orange second
            elif c.severity == ConflictSeverity.WARNING:
                color_order = 2  # Yellow third
            else:
                color_order = 3  # Info/other last
            return (color_order, c.code or "")

        conflicts = sorted(conflicts, key=conflict_sort_key)

        # Add conflict rows
        for row_num, conflict in enumerate(conflicts[:self.config.max_rows], 2):
            # Translate values
            conflict_type_es = translate(conflict.conflict_type.value)
            severity_es = translate(conflict.severity.value)
            message_es = translate(conflict.message)

            ws.cell(row=row_num, column=1, value=conflict.code or "")
            ws.cell(row=row_num, column=2, value=conflict.element_name)
            ws.cell(row=row_num, column=3, value=conflict_type_es)
            ws.cell(row=row_num, column=4, value=severity_es)
            ws.cell(row=row_num, column=5, value=conflict.property_name or "-")
            ws.cell(row=row_num, column=6, value=str(conflict.ifc_value) if conflict.ifc_value is not None else "-")
            ws.cell(row=row_num, column=7, value=str(conflict.bc3_value) if conflict.bc3_value is not None else "-")
            ws.cell(row=row_num, column=8, value=message_es)

            # Color by severity and conflict type (code mismatch = orange)
            fill_color = self._get_severity_color(conflict.severity, conflict.conflict_type)
            for col in range(1, 9):
                self._apply_cell_style(ws.cell(row=row_num, column=col), fill_color)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_matches_sheet(
        self,
        wb: Workbook,
        match_result: MatchResult,
        comparison_result: ComparisonResult,
        sheet_name: str = "Elementos Emparejados"
    ) -> None:
        """Create the matched elements sheet."""
        ws = wb.create_sheet(sheet_name)

        # Headers in Spanish
        headers = [
            ("Codigo/Tag", 15),
            ("Nombre en IFC", 35),
            ("Descripcion en BC3", 40),
            ("Metodo de Emparejamiento", 25),
            ("Estado", 20),
            ("Num. Discrepancias", 18)
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 25

        # Build conflict count map
        conflict_counts = {}
        for conflict in comparison_result.conflicts:
            if conflict.code:
                conflict_counts[conflict.code] = conflict_counts.get(conflict.code, 0) + 1

        # Add matched rows
        for row_num, pair in enumerate(match_result.matched[:self.config.max_rows], 2):
            code = pair.code or ""
            ifc_name = pair.ifc_type.name if pair.ifc_type else ""
            bc3_desc = pair.bc3_element.description if pair.bc3_element else ""
            method = translate(pair.method.value)
            num_conflicts = conflict_counts.get(code, 0)

            if num_conflicts == 0:
                status = "Correcto"
                fill_color = self.config.color_ok
            else:
                status = f"{num_conflicts} discrepancia(s)"
                fill_color = self.config.color_error

            ws.cell(row=row_num, column=1, value=code)
            ws.cell(row=row_num, column=2, value=ifc_name)
            ws.cell(row=row_num, column=3, value=bc3_desc)
            ws.cell(row=row_num, column=4, value=method)
            ws.cell(row=row_num, column=5, value=status)
            ws.cell(row=row_num, column=6, value=num_conflicts)

            for col in range(1, 7):
                self._apply_cell_style(ws.cell(row=row_num, column=col), fill_color)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_missing_sheet(
        self,
        wb: Workbook,
        match_result: MatchResult,
        sheet_name: str,
        status: MatchStatus
    ) -> None:
        """Create a sheet for missing elements."""
        ws = wb.create_sheet(sheet_name)

        if status == MatchStatus.IFC_ONLY:
            items = match_result.ifc_only
            headers = [
                ("Codigo/Tag", 15),
                ("Nombre", 35),
                ("Clase IFC", 20),
                ("Familia", 25),
                ("Tipo", 25),
                ("Accion Requerida", 40)
            ]
            action_text = "Incluir en presupuesto BC3"
        else:
            items = match_result.bc3_only
            headers = [
                ("Codigo", 15),
                ("Descripcion", 40),
                ("Unidad", 10),
                ("Precio", 15),
                ("Familia", 25),
                ("Accion Requerida", 40)
            ]
            action_text = "Modelar en IFC o eliminar del presupuesto"

        # Headers
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 25

        # Add rows
        for row_num, pair in enumerate(items[:self.config.max_rows], 2):
            if status == MatchStatus.IFC_ONLY and pair.ifc_type:
                ifc = pair.ifc_type
                ws.cell(row=row_num, column=1, value=ifc.tag or ifc.global_id)
                ws.cell(row=row_num, column=2, value=ifc.name)
                ws.cell(row=row_num, column=3, value=ifc.ifc_class)
                ws.cell(row=row_num, column=4, value=ifc.family_name or "-")
                ws.cell(row=row_num, column=5, value=ifc.type_name or "-")
                ws.cell(row=row_num, column=6, value=action_text)
            elif status == MatchStatus.BC3_ONLY and pair.bc3_element:
                bc3 = pair.bc3_element
                ws.cell(row=row_num, column=1, value=bc3.code)
                ws.cell(row=row_num, column=2, value=bc3.description)
                ws.cell(row=row_num, column=3, value=bc3.unit)
                ws.cell(row=row_num, column=4, value=f"{bc3.price:.2f} EUR" if bc3.price else "-")
                ws.cell(row=row_num, column=5, value=bc3.family_name or "-")
                ws.cell(row=row_num, column=6, value=action_text)

            # Warning color for all missing items
            for col in range(1, len(headers) + 1):
                self._apply_cell_style(ws.cell(row=row_num, column=col), self.config.color_warning)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _collect_all_properties(self, match_result: MatchResult) -> list[str]:
        """
        Collect all unique property names from all elements.

        Returns:
            Sorted list of unique property names
        """
        property_names = set()

        # From matched elements
        for pair in match_result.matched:
            if pair.ifc_type and pair.ifc_type.properties:
                property_names.update(pair.ifc_type.properties.keys())
            if pair.bc3_element and pair.bc3_element.properties:
                property_names.update(pair.bc3_element.properties.keys())

        # From IFC-only elements
        for pair in match_result.ifc_only:
            if pair.ifc_type and pair.ifc_type.properties:
                property_names.update(pair.ifc_type.properties.keys())

        # From BC3-only elements
        for pair in match_result.bc3_only:
            if pair.bc3_element and pair.bc3_element.properties:
                property_names.update(pair.bc3_element.properties.keys())

        return sorted(property_names)

    def _get_element_status(
        self,
        match_status: MatchStatus,
        codes: list[str],
        comparison_result: 'ComparisonResult'
    ) -> tuple[str, str]:
        """
        Get status text and color for an element group based on conflicts.

        Args:
            match_status: The match status (MATCHED, IFC_ONLY, BC3_ONLY)
            codes: List of codes for elements in this group
            comparison_result: The comparison result with conflicts

        Returns:
            Tuple of (status_text, color_code)
        """
        if match_status == MatchStatus.IFC_ONLY:
            return ("Sin presupuestar", self.config.color_warning)
        elif match_status == MatchStatus.BC3_ONLY:
            return ("Sin modelar", self.config.color_warning)

        # For matched elements, check conflicts across all codes in the group
        total_errors = 0
        total_warnings = 0

        for code in codes:
            for conflict in comparison_result.conflicts:
                if conflict.code == code:
                    if conflict.severity == ConflictSeverity.ERROR:
                        total_errors += 1
                    elif conflict.severity == ConflictSeverity.WARNING:
                        total_warnings += 1

        if total_errors > 0:
            return (f"{total_errors} error(es), {total_warnings} aviso(s)", self.config.color_error)
        elif total_warnings > 0:
            return (f"{total_warnings} aviso(s)", self.config.color_warning)
        else:
            return ("Correcto", self.config.color_ok)

    def _build_dynamic_headers(self, all_properties: list[str]) -> list[tuple[str, int]]:
        """
        Build the header list with fixed columns + dynamic properties.

        Returns:
            List of (header_name, column_width) tuples
        """
        fixed_headers = [
            ("Familia", 25),
            ("Nombre/Descripcion", 40),
            ("Clase IFC", 20),
            ("Tipo", 25),
            ("Unidad", 10),
            ("Cantidad", 12),
            ("Precio (EUR)", 15),
            ("Estado", 30),
            ("Origen", 15),
        ]

        property_headers = [(prop, 18) for prop in all_properties]

        return fixed_headers + property_headers

    def _prepare_element_rows(
        self,
        match_result: MatchResult,
        comparison_result: 'ComparisonResult',
        all_properties: list[str]
    ) -> list[dict]:
        """
        Prepare all element rows, grouping identical elements.

        Elements are grouped by: family + type + ifc_class + all properties.
        Identical elements are merged into a single row with summed quantities.

        Returns:
            List of row dictionaries sorted by family_name
        """
        # Dictionary to group identical elements
        # Key: tuple of (family, name, ifc_class, type_name, unit, price, origin, frozen_properties)
        groups: dict[tuple, dict] = {}

        def process_pair(pair: MatchedPair, status: MatchStatus):
            ifc = pair.ifc_type
            bc3 = pair.bc3_element

            # Extract data with IFC priority
            family = (ifc.family_name if ifc and ifc.family_name else
                     (bc3.family_name if bc3 and bc3.family_name else None))

            name = ifc.name if ifc else (bc3.description if bc3 else "")

            ifc_class = ifc.ifc_class if ifc else ""

            type_name = (ifc.type_name if ifc and ifc.type_name else
                        (bc3.type_name if bc3 and bc3.type_name else None))

            unit = bc3.unit if bc3 else ""

            price = bc3.price if bc3 else None

            # Get quantity (instance_count for IFC, quantity for BC3)
            quantity = 1
            if ifc and hasattr(ifc, 'instance_count') and ifc.instance_count:
                quantity = ifc.instance_count
            elif bc3 and hasattr(bc3, 'quantity') and bc3.quantity:
                quantity = bc3.quantity

            # Origin mapping
            origin_map = {
                MatchStatus.MATCHED: "Emparejado",
                MatchStatus.IFC_ONLY: "Solo IFC",
                MatchStatus.BC3_ONLY: "Solo BC3"
            }
            origin = origin_map.get(status, "")

            # Collect properties (IFC priority)
            properties_values = {}
            for prop_name in all_properties:
                value = None
                if ifc and ifc.properties and prop_name in ifc.properties:
                    value = ifc.properties[prop_name]
                elif bc3 and bc3.properties and prop_name in bc3.properties:
                    value = bc3.properties[prop_name]
                properties_values[prop_name] = value

            # Create grouping key (frozen properties for hashability)
            frozen_props = tuple(sorted(
                (k, str(v) if v is not None else None)
                for k, v in properties_values.items()
            ))

            group_key = (
                family or "",
                name,
                ifc_class,
                type_name or "",
                unit,
                price,
                origin,
                frozen_props
            )

            # Get code for conflict tracking
            code = pair.code or ""

            if group_key in groups:
                groups[group_key]['quantity'] += quantity
                groups[group_key]['codes'].append(code)
            else:
                groups[group_key] = {
                    'family': family,
                    'name': name,
                    'ifc_class': ifc_class,
                    'type_name': type_name,
                    'unit': unit,
                    'quantity': quantity,
                    'price': price,
                    'origin': origin,
                    'match_status': status,
                    'properties': properties_values,
                    'codes': [code]
                }

        # Process all pairs
        for pair in match_result.matched:
            process_pair(pair, MatchStatus.MATCHED)

        for pair in match_result.ifc_only:
            process_pair(pair, MatchStatus.IFC_ONLY)

        for pair in match_result.bc3_only:
            process_pair(pair, MatchStatus.BC3_ONLY)

        # Convert groups to list and add status
        rows = []
        for group_data in groups.values():
            status_text, status_color = self._get_element_status(
                group_data['match_status'],
                group_data['codes'],
                comparison_result
            )
            group_data['status_text'] = status_text
            group_data['status_color'] = status_color
            rows.append(group_data)

        # Sort by family (None at end), then by name
        rows.sort(key=lambda r: (
            r['family'] or 'ZZZZZ',
            r['name'] or ''
        ))

        return rows

    def _create_elements_summary_sheet(
        self,
        wb: Workbook,
        match_result: MatchResult,
        comparison_result: 'ComparisonResult'
    ) -> None:
        """
        Create an element summary sheet with all elements grouped by family.

        This sheet shows ALL elements (matched, IFC-only, BC3-only) with
        dynamic property columns, status indicators, and grouped identical elements.
        """
        ws = wb.create_sheet("Resumen de Elementos")

        # Step 1: Collect all unique properties
        all_properties = self._collect_all_properties(match_result)

        # Step 2: Build headers
        headers = self._build_dynamic_headers(all_properties)

        # Step 3: Write headers
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 25

        # Step 4: Prepare grouped rows
        rows_data = self._prepare_element_rows(match_result, comparison_result, all_properties)

        # Step 5: Write data rows
        for row_num, row_data in enumerate(rows_data[:self.config.max_rows], 2):
            col = 1

            # Fixed columns
            ws.cell(row=row_num, column=col, value=row_data['family'] or "-")
            col += 1
            ws.cell(row=row_num, column=col, value=row_data['name'])
            col += 1
            ws.cell(row=row_num, column=col, value=row_data['ifc_class'] or "-")
            col += 1
            ws.cell(row=row_num, column=col, value=row_data['type_name'] or "-")
            col += 1
            ws.cell(row=row_num, column=col, value=row_data['unit'] or "-")
            col += 1
            ws.cell(row=row_num, column=col, value=row_data['quantity'] or 0)
            col += 1

            # Price formatted
            price = row_data['price']
            ws.cell(row=row_num, column=col, value=f"{price:.2f}" if price else "-")
            col += 1

            # Status
            ws.cell(row=row_num, column=col, value=row_data['status_text'])
            col += 1

            # Origin
            ws.cell(row=row_num, column=col, value=row_data['origin'])
            col += 1

            # Dynamic property columns
            for prop_name in all_properties:
                prop_value = row_data['properties'].get(prop_name)
                ws.cell(row=row_num, column=col,
                       value=str(prop_value) if prop_value is not None else "-")
                col += 1

            # Apply row styling with status color
            fill_color = row_data['status_color']
            for c in range(1, col):
                self._apply_cell_style(ws.cell(row=row_num, column=c), fill_color)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _get_severity_color(self, severity: ConflictSeverity, conflict_type: ConflictType = None) -> str:
        """Get the color for a severity level and conflict type."""
        # Special case: code mismatch gets orange
        if conflict_type == ConflictType.CODE_MISMATCH:
            return self.config.color_code_mismatch

        if severity == ConflictSeverity.ERROR:
            return self.config.color_error
        elif severity == ConflictSeverity.WARNING:
            return self.config.color_warning
        elif severity == ConflictSeverity.INFO:
            return self.config.color_info
        return "FFFFFF"

    def generate_json_report(
        self,
        match_result: MatchResult,
        comparison_result: ComparisonResult
    ) -> dict:
        """
        Generate a JSON-serializable report.

        Args:
            match_result: Result from the Matcher
            comparison_result: Result from the Comparator

        Returns:
            Dictionary with report data
        """
        return {
            "resumen": {
                "fecha_generacion": datetime.now().isoformat(),
                "emparejamiento": {
                    "total_tipos_ifc": match_result.total_ifc_types,
                    "total_elementos_bc3": match_result.total_bc3_elements,
                    "emparejados": len(match_result.matched),
                    "solo_ifc": len(match_result.ifc_only),
                    "solo_bc3": len(match_result.bc3_only),
                    "tasa_emparejamiento": f"{match_result.match_rate:.1f}%"
                },
                "comparacion": comparison_result.summary()
            },
            "discrepancias": [
                {
                    "codigo": c.code,
                    "elemento": c.element_name,
                    "tipo": translate(c.conflict_type.value),
                    "gravedad": translate(c.severity.value),
                    "propiedad": c.property_name,
                    "valor_ifc": str(c.ifc_value) if c.ifc_value is not None else None,
                    "valor_bc3": str(c.bc3_value) if c.bc3_value is not None else None,
                    "mensaje": translate(c.message)
                }
                for c in comparison_result.conflicts
            ],
            "emparejados": [
                {
                    "codigo": p.code,
                    "metodo": translate(p.method.value),
                    "confianza": p.confidence
                }
                for p in match_result.matched
            ],
            "solo_ifc": [
                {"codigo": p.code, "nombre": p.name}
                for p in match_result.ifc_only
            ],
            "solo_bc3": [
                {"codigo": p.code, "nombre": p.name}
                for p in match_result.bc3_only
            ]
        }
