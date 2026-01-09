import ifcopenshell
from ifcopenshell.util.element import get_psets
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os


def parse_bc3(filepath):
    data = {}
    with open(filepath, 'r', encoding='latin1', errors='replace') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()
        if line.startswith('~C|'):
            parts = line.split('|')
            if len(parts) > 2:
                code = parts[1].replace('\\', '').strip()
                unit = parts[2].strip()
                desc = parts[3].strip() if len(parts) > 3 else "Sense descripció"
                data[code] = {'desc': desc, 'unit': unit, 'qty': 0.0}

    for line in lines:
        if line.startswith('~M|'):
            parts = line.split('|')
            if len(parts) > 2:
                code = parts[1].strip()
                if code in data:
                    for part in reversed(parts):
                        try:
                            val = float(part.replace(',', '.'))
                            data[code]['qty'] += val
                            break
                        except ValueError:
                            continue
    return data


def parse_ifc(filepath):
    model = ifcopenshell.open(filepath)
    elements = model.by_type("IfcElement")

    result = {}
    for el in elements:
        code = el.Name or el.GlobalId
        props = get_psets(el)

        code_props = ["Assembly Code", "Reference", "Codi", "Code", "Partida"]
        for pset in props.values():
            for k, v in pset.items():
                if k in code_props and v:
                    code = str(v).strip()
                    break

        qty = 1.0
        unit = "u"
        for pset in props.values():
            for k, v in pset.items():
                if isinstance(v, (int, float)):
                    if "Volume" in k or "Volumen" in k:
                        qty = v; unit = "m3"; break
                    elif "Area" in k:
                        qty = v; unit = "m2"; break
                    elif "Length" in k or "Longitud" in k:
                        qty = v; unit = "m"; break

        if code not in result:
            result[code] = {'qty': 0.0, 'unit': unit}

        result[code]['qty'] += qty
        result[code]['unit'] = unit
    return result


def generate_basic_discrepancy_report(ifc_path: str, bc3_path: str, output_excel: str, tolerance: float = 0.1):
    """
    Compara dades d'un arxiu IFC i un BC3, i genera un Excel amb discrepàncies de Codi, Unitat i Quantitat.
    Marca en vermell els errors greus, i en groc els codis que només apareixen en un dels arxius.
    """
    ifc_data = parse_ifc(ifc_path)
    bc3_data = parse_bc3(bc3_path)

    codis_ifc = set(ifc_data.keys())
    codis_bc3 = set(bc3_data.keys())

    discrepancies = []

    # 🔴 Discrepàncies greus
    for code in sorted(codis_ifc & codis_bc3):
        ifc_item = ifc_data[code]
        bc3_item = bc3_data[code]

        unit_match = ifc_item['unit'].strip().lower() == bc3_item['unit'].strip().lower()
        qty_diff = abs(ifc_item['qty'] - bc3_item['qty'])
        qty_match = qty_diff <= tolerance

        if not unit_match or not qty_match:
            tipus = []
            if not unit_match:
                tipus.append("Unitat diferent")
            if not qty_match:
                tipus.append("Quantitat diferent")

            discrepancies.append({
                "Codi": code,
                "Descripció": bc3_item['desc'],
                "Tipus de discrepància": ", ".join(tipus),
                "Valor IFC": f"{ifc_item['qty']:.2f} {ifc_item['unit']}",
                "Valor BC3": f"{bc3_item['qty']:.2f} {bc3_item['unit']}",
                "Color": "vermell"
            })

    # 🟡 Només a l'IFC
    for code in sorted(codis_ifc - codis_bc3):
        item = ifc_data[code]
        discrepancies.append({
            "Codi": code,
            "Descripció": "(Només al model IFC)",
            "Tipus de discrepància": "No pressupostat",
            "Valor IFC": f"{item['qty']:.2f} {item['unit']}",
            "Valor BC3": "-",
            "Color": "groc"
        })

    # 🟡 Només al BC3
    for code in sorted(codis_bc3 - codis_ifc):
        item = bc3_data[code]
        discrepancies.append({
            "Codi": code,
            "Descripció": item['desc'],
            "Tipus de discrepància": "No modelat",
            "Valor IFC": "-",
            "Valor BC3": f"{item['qty']:.2f} {item['unit']}",
            "Color": "groc"
        })

    # Ordenar: vermells primer
    discrepancies = sorted(discrepancies, key=lambda x: 0 if x['Color'] == "vermell" else 1)

    # 📄 Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepàncies"

    headers = ["Codi", "Descripció", "Tipus de discrepància", "Valor IFC", "Valor BC3"]
    ws.append(headers)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C6500")

    for item in discrepancies:
        row = [item[h] for h in headers]
        ws.append(row)

        for col in range(1, len(headers)+1):
            if item["Color"] == "vermell":
                ws.cell(row=ws.max_row, column=col).fill = red_fill
                ws.cell(row=ws.max_row, column=col).font = red_font
            elif item["Color"] == "groc":
                ws.cell(row=ws.max_row, column=col).fill = yellow_fill
                ws.cell(row=ws.max_row, column=col).font = yellow_font

    wb.save(output_excel)
    print(f"\n✅ Informe generat correctament: {output_excel}")
