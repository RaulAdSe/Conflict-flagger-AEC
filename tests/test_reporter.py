"""Tests for Reporter."""

import pytest
from pathlib import Path
import tempfile
import json

from src.parsers.bc3_parser import BC3Element
from src.parsers.ifc_parser import IFCType
from src.matching.matcher import MatchedPair, MatchResult, MatchStatus, MatchMethod
from src.comparison.comparator import (
    ComparisonResult, Conflict, ConflictType, ConflictSeverity
)
from src.reporting.reporter import Reporter, ReportConfig


class TestReportConfig:
    """Tests for ReportConfig dataclass."""

    def test_default_config(self):
        """Test default configuration values."""
        config = ReportConfig()
        assert config.color_error == "FF9999"
        assert config.color_warning == "FFFF99"
        assert config.color_ok == "99FF99"
        assert config.show_ok_matches is True

    def test_custom_config(self):
        """Test custom configuration."""
        config = ReportConfig(
            color_error="FF0000",
            show_ok_matches=False,
            max_rows=100
        )
        assert config.color_error == "FF0000"
        assert config.show_ok_matches is False
        assert config.max_rows == 100


class TestReporter:
    """Tests for Reporter class."""

    @pytest.fixture
    def reporter(self):
        """Create a reporter instance."""
        return Reporter()

    @pytest.fixture
    def sample_match_result(self):
        """Create sample match result."""
        bc3_1 = BC3Element(
            code="350147",
            unit="m3",
            description="Column 600x600",
            price=150.0,
            family_name="Pilar",
            type_name="600x600"
        )
        ifc_1 = IFCType(
            global_id="guid1",
            tag="350147",
            name="Pilar:600x600",
            ifc_class="IfcColumnType",
            family_name="Pilar",
            type_name="600x600"
        )

        matched_pair = MatchedPair(
            status=MatchStatus.MATCHED,
            method=MatchMethod.TAG,
            ifc_type=ifc_1,
            bc3_element=bc3_1,
            match_key="350147"
        )

        ifc_only = IFCType(
            global_id="guid2",
            tag="999",
            name="Unbudgeted:Type",
            ifc_class="IfcWallType"
        )
        ifc_only_pair = MatchedPair(
            status=MatchStatus.IFC_ONLY,
            method=MatchMethod.NONE,
            ifc_type=ifc_only
        )

        bc3_only = BC3Element(
            code="orphan",
            unit="m2",
            description="Orphan item",
            price=50.0
        )
        bc3_only_pair = MatchedPair(
            status=MatchStatus.BC3_ONLY,
            method=MatchMethod.NONE,
            bc3_element=bc3_only
        )

        return MatchResult(
            matched=[matched_pair],
            ifc_only=[ifc_only_pair],
            bc3_only=[bc3_only_pair],
            total_ifc_types=2,
            total_bc3_elements=2,
            match_count=1
        )

    @pytest.fixture
    def sample_comparison_result(self):
        """Create sample comparison result."""
        conflicts = [
            Conflict(
                conflict_type=ConflictType.PROPERTY_MISMATCH,
                severity=ConflictSeverity.ERROR,
                code="350147",
                element_name="Column 600x600",
                property_name="h",
                ifc_value=0.8,
                bc3_value=0.6,
                message="Height differs"
            ),
            Conflict(
                conflict_type=ConflictType.MISSING_IN_BC3,
                severity=ConflictSeverity.WARNING,
                code="999",
                element_name="Unbudgeted:Type",
                message="Element exists in IFC but not in BC3"
            ),
            Conflict(
                conflict_type=ConflictType.MISSING_IN_IFC,
                severity=ConflictSeverity.WARNING,
                code="orphan",
                element_name="Orphan item",
                message="Element exists in BC3 but not in IFC"
            ),
        ]

        return ComparisonResult(
            conflicts=conflicts,
            missing_in_bc3=1,
            missing_in_ifc=1,
            property_mismatches=1,
            total_matched=1,
            total_with_conflicts=3
        )

    def test_generate_excel_report(self, reporter, sample_match_result, sample_comparison_result):
        """Test generating an Excel report."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"

            result_path = reporter.generate_report(
                sample_match_result,
                sample_comparison_result,
                output_path
            )

            assert result_path.exists()
            assert result_path.suffix == ".xlsx"

    def test_report_has_expected_sheets(self, reporter, sample_match_result, sample_comparison_result):
        """Test that report has all expected sheets."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"

            reporter.generate_report(
                sample_match_result,
                sample_comparison_result,
                output_path
            )

            wb = load_workbook(output_path)
            sheet_names = wb.sheetnames

            assert "Summary" in sheet_names
            assert "Conflicts" in sheet_names
            assert "Matched Elements" in sheet_names
            assert "Missing in BC3" in sheet_names
            assert "Missing in IFC" in sheet_names

    def test_summary_sheet_content(self, reporter, sample_match_result, sample_comparison_result):
        """Test summary sheet has correct content."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"

            reporter.generate_report(
                sample_match_result,
                sample_comparison_result,
                output_path
            )

            wb = load_workbook(output_path)
            ws = wb["Summary"]

            # Check title
            assert "IFC-BC3 Comparison Report" in ws["A1"].value

    def test_conflicts_sheet_has_data(self, reporter, sample_match_result, sample_comparison_result):
        """Test conflicts sheet has data rows."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"

            reporter.generate_report(
                sample_match_result,
                sample_comparison_result,
                output_path
            )

            wb = load_workbook(output_path)
            ws = wb["Conflicts"]

            # Should have header row + 3 conflict rows
            assert ws.max_row >= 4

    def test_generate_json_report(self, reporter, sample_match_result, sample_comparison_result):
        """Test generating a JSON report."""
        report = reporter.generate_json_report(
            sample_match_result,
            sample_comparison_result
        )

        assert "summary" in report
        assert "conflicts" in report
        assert "matched" in report
        assert "ifc_only" in report
        assert "bc3_only" in report

        # Check conflicts
        assert len(report["conflicts"]) == 3

        # Check matched
        assert len(report["matched"]) == 1
        assert report["matched"][0]["code"] == "350147"

    def test_json_report_is_serializable(self, reporter, sample_match_result, sample_comparison_result):
        """Test that JSON report can be serialized."""
        report = reporter.generate_json_report(
            sample_match_result,
            sample_comparison_result
        )

        # Should not raise
        json_str = json.dumps(report)
        assert len(json_str) > 0

    def test_custom_config(self, sample_match_result, sample_comparison_result):
        """Test reporter with custom config."""
        config = ReportConfig(
            color_error="FF0000",
            show_ok_matches=False,
            max_rows=1
        )
        reporter = Reporter(config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"

            result_path = reporter.generate_report(
                sample_match_result,
                sample_comparison_result,
                output_path
            )

            assert result_path.exists()

    def test_empty_results(self, reporter):
        """Test with empty results."""
        match_result = MatchResult(
            matched=[],
            ifc_only=[],
            bc3_only=[],
            total_ifc_types=0,
            total_bc3_elements=0,
            match_count=0
        )

        comparison_result = ComparisonResult(
            conflicts=[],
            missing_in_bc3=0,
            missing_in_ifc=0,
            property_mismatches=0,
            total_matched=0
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "empty_report.xlsx"

            result_path = reporter.generate_report(
                match_result,
                comparison_result,
                output_path
            )

            assert result_path.exists()


class TestReporterWithRealFiles:
    """Tests using real IFC and BC3 files."""

    @pytest.fixture
    def real_ifc_path(self):
        path = Path(__file__).parent.parent / "data" / "input" / "GUIA MODELADO V2.ifc"
        if not path.exists():
            pytest.skip(f"Real IFC file not found: {path}")
        return path

    @pytest.fixture
    def real_bc3_path(self):
        path = Path(__file__).parent.parent / "data" / "input" / "GUIA MODELADO V2 2025-12-18 06-47-01.bc3"
        if not path.exists():
            pytest.skip(f"Real BC3 file not found: {path}")
        return path

    def test_generate_report_from_real_files(self, real_ifc_path, real_bc3_path):
        """Test generating report from real files."""
        from src.parsers.ifc_parser import IFCParser
        from src.parsers.bc3_parser import BC3Parser
        from src.matching.matcher import Matcher
        from src.comparison.comparator import Comparator

        ifc_parser = IFCParser()
        bc3_parser = BC3Parser()
        matcher = Matcher()
        comparator = Comparator()
        reporter = Reporter()

        ifc_result = ifc_parser.parse(real_ifc_path)
        bc3_result = bc3_parser.parse(real_bc3_path)
        match_result = matcher.match(ifc_result, bc3_result)
        comparison = comparator.compare(match_result)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "real_report.xlsx"

            result_path = reporter.generate_report(
                match_result,
                comparison,
                output_path
            )

            assert result_path.exists()
            print(f"\nReport generated: {result_path}")
            print(f"  Matched: {len(match_result.matched)}")
            print(f"  Conflicts: {len(comparison.conflicts)}")
