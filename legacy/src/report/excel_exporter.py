import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def generate_excel_report(report_df, output_path):
    """
    Writes the dataframe to Excel and applies conditional formatting.
    """
    # Create basic Excel
    # Drop 'Color' column for the export, use it for formatting later
    data_to_write = report_df.drop(columns=['Color'])
    data_to_write.to_excel(output_path, index=False, sheet_name='Matriz Control')
    
    # Load for formatting
    wb = load_workbook(output_path)
    ws = wb['Matriz Control']
    
    # Define Fills
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Iterate rows 
    # (openpyxl 1-indexed, header is row 1)
    for idx, row in enumerate(report_df.itertuples(), start=2):
        color_code = row.Color
        
        # Apply to the CHECK column (last one usually, verify index)
        # Columns: Familia, Subgrupo, Codi, Variable, Revit, Presto, Memoria, CHECK
        # CHECK is column 8 (H)
        
        cell = ws.cell(row=idx, column=8)
        
        if color_code == 'RED':
            cell.fill = red_fill
        elif color_code == 'GREEN':
            cell.fill = green_fill
        elif color_code == 'YELLOW':
            cell.fill = yellow_fill
            
    wb.save(output_path)
    print(f"Report saved to {output_path}")
